import openpyxl

# wokrbook = openpyxl.load_workbook("data.xlsx")
# sheet = wokrbook.active
# print(sheet["A1"].value)
# sheet["A1"].value = "CID"
# wokrbook.save("data.xlsx")

my_file = openpyxl.load_workbook("tabelka.xlsx")
sheet = my_file.active
qualifica = []
lenght = sheet.max_row
for i in range(1, lenght):
    qualifica.append(sheet.cell(row=i, column=1).value)

print(qualifica)


def get_a_column(sheet_name, nr_of_column):
    values_from_column = []
    lenght = sheet.max_row
    for i in range(1, lenght):
        values_from_column.append(sheet.cell(row=i, column=nr_of_column).value)
    return values_from_column


procent = get_a_column(sheet, 2)
print(procent)

kod = get_a_column(sheet, 3)
print(kod)

różnica = get_a_column(sheet, 4)
# print(f'{różnica}:0.2f')

indexes_qualifica_not_d_z = []
for i in range(2, len(qualifica)):
    if qualifica[i] != 'D' and qualifica[i] != 'Z':
        indexes_qualifica_not_d_z.append(i)

print(indexes_qualifica_not_d_z)

indexes_procent_not_100 = []
for i in range(2, len(procent)):
    if procent[i] != 100:
        indexes_procent_not_100.append(i)

print(indexes_procent_not_100)

set_indexes_qualifica = set(indexes_qualifica_not_d_z)
set_indexes_procens = set(indexes_procent_not_100)
both = set_indexes_procens & set_indexes_qualifica
print(both)

difference = []
for element in both:
    if różnica[element] != None:
        difference.append(różnica[element])

print(difference)
difference_sum = sum(difference)
print(f'{sum(difference):0.2f}')
sheets = my_file.sheetnames

if "Scores" not in sheets:
    my_file.create_sheet("Scores")

scores = my_file["Scores"]
scores['A1'] = "Not D and Z, part time"
scores['A2'] = f'{difference_sum:0.2f}'
my_file.save('tabelka.xlsx')
